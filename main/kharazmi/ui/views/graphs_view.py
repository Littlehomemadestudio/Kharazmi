"""
GraphsView — full-page visual analytics for RASK! routes/workflows.

Features:
  - Landing page: select a workflow (from current route or journal history)
  - CSV / Excel import/export for workflows
  - Multiple chart types: success probability bar chart, risk heatmap,
    duration timeline, branch flow diagram, step kind distribution
  - Beautiful gold-on-dark theme with animated transitions
"""
from __future__ import annotations

import csv
import io
import json
import os
from pathlib import Path
from typing import Optional

from PySide6.QtCore import Qt, QRectF, Signal, QTimer, QPropertyAnimation, QEasingCurve
from PySide6.QtGui import (
    QPainter, QPainterPath, QColor, QPen, QBrush, QFont,
    QLinearGradient, QRadialGradient, QConicalGradient, QFontMetrics,
)
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame,
    QStackedWidget, QScrollArea, QFileDialog, QGridLayout, QComboBox,
    QSizePolicy, QMessageBox, QToolButton, QGroupBox,
)

from ...ai import Route, RouteStep, RouteEdge, Insight, JournalStore
from ..theme import Palette


# ──────────────────────────────────────────────────────────────────────
#  Workflow CSV/Excel I/O
# ──────────────────────────────────────────────────────────────────────

def export_route_csv(route: Route, path: str) -> str:
    """Export a Route to CSV (steps + edges in one file)."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        # Steps section
        w.writerow(["=== STEPS ==="])
        w.writerow([
            "id", "title", "description", "duration_minutes", "success_probability",
            "location", "fallback", "depends_on", "sub_goals", "cost_estimate",
            "risk_level", "branch", "kind", "x", "y",
        ])
        for s in route.steps:
            w.writerow([
                s.id, s.title, s.description, s.duration_minutes,
                f"{s.success_probability:.2f}", s.location, s.fallback,
                ";".join(s.depends_on), ";".join(s.sub_goals),
                s.cost_estimate, s.risk_level, s.branch, s.kind,
                f"{s.x_hint:.1f}", f"{s.y_hint:.1f}",
            ])
        w.writerow([])
        # Edges section
        w.writerow(["=== EDGES ==="])
        w.writerow(["source_id", "target_id", "kind", "label"])
        for e in route.edges:
            w.writerow([e.source_id, e.target_id, e.kind, e.label])
        w.writerow([])
        # Insights section
        w.writerow(["=== INSIGHTS ==="])
        w.writerow(["kind", "title", "body", "anchor_step_id", "x_hint", "y_hint"])
        for i in route.insights:
            w.writerow([i.kind, i.title, i.body, i.anchor_step_id or "",
                        f"{i.x_hint:.2f}", f"{i.y_hint:.2f}"])
        w.writerow([])
        # Route metadata
        w.writerow(["=== METADATA ==="])
        w.writerow(["goal", route.goal])
        w.writerow(["summary", route.summary])
        w.writerow(["overall_success_probability", f"{route.overall_success_probability:.2f}"])
        w.writerow(["total_duration_minutes", route.total_duration_minutes])
        w.writerow(["layout_style", route.layout_style])
    return str(p)


def export_route_excel(route: Route, path: str) -> str:
    """Export a Route to Excel (.xlsx) with multiple sheets."""
    import xlsxwriter
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = xlsxwriter.Workbook(str(p))

    # Formats
    hdr_fmt = wb.add_format({"bold": True, "bg_color": Palette.GOLD_DEEP, "color": Palette.TEXT_ON_GOLD, "border": 1})
    cell_fmt = wb.add_format({"border": 1, "text_wrap": True})

    # Steps sheet
    ws = wb.add_worksheet("Steps")
    step_headers = ["ID", "Title", "Description", "Duration (min)", "Success Prob",
                    "Location", "Fallback", "Depends On", "Sub Goals", "Cost",
                    "Risk Level", "Branch", "Kind", "X", "Y"]
    for col, h in enumerate(step_headers):
        ws.write(0, col, h, hdr_fmt)
    for row, s in enumerate(route.steps, 1):
        ws.write(row, 0, s.id, cell_fmt)
        ws.write(row, 1, s.title, cell_fmt)
        ws.write(row, 2, s.description, cell_fmt)
        ws.write(row, 3, s.duration_minutes, cell_fmt)
        ws.write(row, 4, s.success_probability, cell_fmt)
        ws.write(row, 5, s.location, cell_fmt)
        ws.write(row, 6, s.fallback, cell_fmt)
        ws.write(row, 7, ";".join(s.depends_on), cell_fmt)
        ws.write(row, 8, ";".join(s.sub_goals), cell_fmt)
        ws.write(row, 9, s.cost_estimate, cell_fmt)
        ws.write(row, 10, s.risk_level, cell_fmt)
        ws.write(row, 11, s.branch, cell_fmt)
        ws.write(row, 12, s.kind, cell_fmt)
        ws.write(row, 13, s.x_hint, cell_fmt)
        ws.write(row, 14, s.y_hint, cell_fmt)

    # Edges sheet
    we = wb.add_worksheet("Edges")
    edge_headers = ["Source ID", "Target ID", "Kind", "Label"]
    for col, h in enumerate(edge_headers):
        we.write(0, col, h, hdr_fmt)
    for row, e in enumerate(route.edges, 1):
        we.write(row, 0, e.source_id, cell_fmt)
        we.write(row, 1, e.target_id, cell_fmt)
        we.write(row, 2, e.kind, cell_fmt)
        we.write(row, 3, e.label, cell_fmt)

    # Insights sheet
    wi = wb.add_worksheet("Insights")
    ins_headers = ["Kind", "Title", "Body", "Anchor Step", "X Hint", "Y Hint"]
    for col, h in enumerate(ins_headers):
        wi.write(0, col, h, hdr_fmt)
    for row, ins in enumerate(route.insights, 1):
        wi.write(row, 0, ins.kind, cell_fmt)
        wi.write(row, 1, ins.title, cell_fmt)
        wi.write(row, 2, ins.body, cell_fmt)
        wi.write(row, 3, ins.anchor_step_id or "", cell_fmt)
        wi.write(row, 4, ins.x_hint, cell_fmt)
        wi.write(row, 5, ins.y_hint, cell_fmt)

    # Summary sheet
    wm = wb.add_worksheet("Summary")
    wm.write(0, 0, "Goal", hdr_fmt)
    wm.write(0, 1, route.goal, cell_fmt)
    wm.write(1, 0, "Summary", hdr_fmt)
    wm.write(1, 1, route.summary, cell_fmt)
    wm.write(2, 0, "Overall Success", hdr_fmt)
    wm.write(2, 1, route.overall_success_probability, cell_fmt)
    wm.write(3, 0, "Total Duration (min)", hdr_fmt)
    wm.write(3, 1, route.total_duration_minutes, cell_fmt)
    wm.write(4, 0, "Layout Style", hdr_fmt)
    wm.write(4, 1, route.layout_style, cell_fmt)
    wm.write(5, 0, "Step Count", hdr_fmt)
    wm.write(5, 1, len(route.steps), cell_fmt)
    wm.write(6, 0, "Edge Count", hdr_fmt)
    wm.write(6, 1, len(route.edges), cell_fmt)

    wb.close()
    return str(p)


def import_route_csv(path: str) -> Route:
    """Import a Route from CSV file."""
    p = Path(path)
    text = p.read_text(encoding="utf-8-sig")
    lines = text.strip().split("\n")

    section = None
    steps_data = []
    edges_data = []
    insights_data = []
    metadata = {}

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if stripped == "=== STEPS ===":
            section = "steps"
            continue
        elif stripped == "=== EDGES ===":
            section = "edges"
            continue
        elif stripped == "=== INSIGHTS ===":
            section = "insights"
            continue
        elif stripped == "=== METADATA ===":
            section = "metadata"
            continue

        # Skip header rows
        if section == "steps" and stripped.startswith("id,"):
            continue
        if section == "edges" and stripped.startswith("source_id,"):
            continue
        if section == "insights" and stripped.startswith("kind,"):
            continue

        reader = csv.reader(io.StringIO(stripped))
        row = next(reader, None)
        if not row:
            continue

        if section == "steps":
            steps_data.append(row)
        elif section == "edges":
            edges_data.append(row)
        elif section == "insights":
            insights_data.append(row)
        elif section == "metadata" and len(row) >= 2:
            metadata[row[0]] = row[1]

    # Build Route
    steps = []
    for row in steps_data:
        try:
            step = RouteStep(
                id=row[0] if len(row) > 0 else "",
                title=row[1] if len(row) > 1 else "Untitled",
                description=row[2] if len(row) > 2 else "",
                duration_minutes=int(float(row[3])) if len(row) > 3 and row[3] else 0,
                success_probability=float(row[4]) if len(row) > 4 and row[4] else 0.5,
                location=row[5] if len(row) > 5 else "",
                fallback=row[6] if len(row) > 6 else "",
                depends_on=row[7].split(";") if len(row) > 7 and row[7] else [],
                sub_goals=row[8].split(";") if len(row) > 8 and row[8] else [],
                cost_estimate=row[9] if len(row) > 9 else "",
                risk_level=row[10] if len(row) > 10 else "low",
                branch=row[11] if len(row) > 11 else "main",
                kind=row[12] if len(row) > 12 else "action",
                x_hint=float(row[13]) if len(row) > 13 and row[13] else 0.0,
                y_hint=float(row[14]) if len(row) > 14 and row[14] else 0.0,
            )
            steps.append(step)
        except (ValueError, IndexError):
            continue

    edges = []
    for row in edges_data:
        try:
            edges.append(RouteEdge(
                source_id=row[0] if len(row) > 0 else "",
                target_id=row[1] if len(row) > 1 else "",
                kind=row[2] if len(row) > 2 else "primary",
                label=row[3] if len(row) > 3 else "",
            ))
        except (ValueError, IndexError):
            continue

    insights = []
    for row in insights_data:
        try:
            insights.append(Insight(
                kind=row[0] if len(row) > 0 else "improvement",
                title=row[1] if len(row) > 1 else "",
                body=row[2] if len(row) > 2 else "",
                anchor_step_id=row[3] if len(row) > 3 and row[3] else None,
                x_hint=float(row[4]) if len(row) > 4 and row[4] else 0.5,
                y_hint=float(row[5]) if len(row) > 5 and row[5] else 0.5,
            ))
        except (ValueError, IndexError):
            continue

    return Route(
        goal=metadata.get("goal", "Imported Workflow"),
        steps=steps,
        edges=edges,
        insights=insights,
        overall_success_probability=float(metadata.get("overall_success_probability", 0.5)),
        total_duration_minutes=int(metadata.get("total_duration_minutes", 0)),
        summary=metadata.get("summary", ""),
        layout_style=metadata.get("layout_style", "imported"),
    )


def import_route_excel(path: str) -> Route:
    """Import a Route from Excel (.xlsx) file."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    steps = []
    if "Steps" in wb.sheetnames:
        ws = wb["Steps"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                deps = str(row[7]).split(";") if row[7] else []
                sgs = str(row[8]).split(";") if row[8] else []
                steps.append(RouteStep(
                    id=str(row[0]), title=str(row[1] or "Untitled"),
                    description=str(row[2] or ""),
                    duration_minutes=int(float(row[3] or 0)),
                    success_probability=float(row[4] or 0.5),
                    location=str(row[5] or ""), fallback=str(row[6] or ""),
                    depends_on=deps, sub_goals=sgs,
                    cost_estimate=str(row[9] or ""),
                    risk_level=str(row[10] or "low"),
                    branch=str(row[11] or "main"),
                    kind=str(row[12] or "action"),
                    x_hint=float(row[13] or 0), y_hint=float(row[14] or 0),
                ))
            except (ValueError, IndexError):
                continue

    edges = []
    if "Edges" in wb.sheetnames:
        ws = wb["Edges"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                edges.append(RouteEdge(
                    source_id=str(row[0]), target_id=str(row[1]),
                    kind=str(row[2] or "primary"), label=str(row[3] or ""),
                ))
            except (ValueError, IndexError):
                continue

    insights = []
    if "Insights" in wb.sheetnames:
        ws = wb["Insights"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                insights.append(Insight(
                    kind=str(row[0]), title=str(row[1] or ""),
                    body=str(row[2] or ""),
                    anchor_step_id=str(row[3]) if row[3] else None,
                    x_hint=float(row[4] or 0.5), y_hint=float(row[5] or 0.5),
                ))
            except (ValueError, IndexError):
                continue

    metadata = {}
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                metadata[str(row[0])] = row[1]

    wb.close()
    return Route(
        goal=str(metadata.get("Goal", "Imported Workflow")),
        steps=steps, edges=edges, insights=insights,
        overall_success_probability=float(metadata.get("Overall Success", 0.5)),
        total_duration_minutes=int(metadata.get("Total Duration (min)", 0) or 0),
        summary=str(metadata.get("Summary", "")),
        layout_style=str(metadata.get("Layout Style", "imported")),
    )


# ──────────────────────────────────────────────────────────────────────
#  Chart Widgets (custom QPainter-based)
# ──────────────────────────────────────────────────────────────────────

class _BarChartWidget(QWidget):
    """Horizontal bar chart showing step success probabilities."""

    def __init__(self, route: Route, parent=None):
        super().__init__(parent)
        self.route = route
        self.setMinimumHeight(max(300, len(route.steps) * 38 + 60))
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setRenderHint(QPainter.TextAntialiasing, True)

        w, h = self.width(), self.height()
        steps = self.route.steps
        if not steps:
            p.setPen(QColor(Palette.TEXT_TERTIARY))
            p.setFont(QFont("Inter", 12))
            p.drawText(self.rect(), Qt.AlignCenter, "No steps to display")
            p.end()
            return

        bar_h = min(32, max(18, (h - 50) // len(steps) - 4))
        label_w = 160
        bar_area = w - label_w - 80
        y = 10

        title_font = QFont("Inter", 10, QFont.Bold)
        p.setFont(title_font)
        p.setPen(QColor(Palette.GOLD_BRIGHT))
        p.drawText(10, y + 12, "Success Probability by Step")
        y += 28

        bar_font = QFont("Inter", 9)
        val_font = QFont("JetBrains Mono", 9, QFont.Bold)

        for step in steps:
            # Label
            p.setFont(bar_font)
            p.setPen(QColor(Palette.TEXT_SECONDARY))
            label = step.title[:20] + "…" if len(step.title) > 20 else step.title
            p.drawText(QRectF(4, y, label_w - 8, bar_h), Qt.AlignRight | Qt.AlignVCenter, label)

            # Bar background
            bar_x = label_w
            p.setPen(Qt.NoPen)
            p.setBrush(QColor(Palette.BG_TERTIARY))
            p.drawRoundedRect(QRectF(bar_x, y + 2, bar_area, bar_h - 4), 4, 4)

            # Bar fill
            prob = max(0, min(1, step.success_probability))
            bar_fill_w = bar_area * prob
            if prob > 0.7:
                fill_color = QColor("#5A8A5A")
            elif prob > 0.4:
                fill_color = QColor("#D4AF37")
            else:
                fill_color = QColor("#A85A5A")

            grad = QLinearGradient(bar_x, 0, bar_x + bar_fill_w, 0)
            grad.setColorAt(0, fill_color)
            grad.setColorAt(1, fill_color.lighter(120))
            p.setBrush(QBrush(grad))
            p.drawRoundedRect(QRectF(bar_x, y + 2, bar_fill_w, bar_h - 4), 4, 4)

            # Value text
            p.setFont(val_font)
            p.setPen(QColor(Palette.TEXT_PRIMARY))
            p.drawText(QRectF(bar_x + bar_area + 4, y, 60, bar_h),
                       Qt.AlignLeft | Qt.AlignVCenter, f"{prob:.0%}")

            y += bar_h + 4

        p.end()


class _RiskHeatmapWidget(QWidget):
    """Grid-based heatmap of step risk levels."""

    def __init__(self, route: Route, parent=None):
        super().__init__(parent)
        self.route = route
        self.setFixedHeight(220)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)

        w, h = self.width(), self.height()
        steps = self.route.steps
        if not steps:
            p.end()
            return

        title_font = QFont("Inter", 10, QFont.Bold)
        p.setFont(title_font)
        p.setPen(QColor(Palette.GOLD_BRIGHT))
        p.drawText(10, 18, "Risk Level Distribution")

        risk_colors = {
            "low": QColor("#5A8A5A"),
            "medium": QColor("#D4AF37"),
            "high": QColor("#A87A4A"),
            "severe": QColor("#A85A5A"),
        }

        # Count risks
        risk_counts = {"low": 0, "medium": 0, "high": 0, "severe": 0}
        for s in steps:
            risk_counts[s.risk_level] = risk_counts.get(s.risk_level, 0) + 1

        total = max(1, len(steps))
        y = 36
        bar_area_w = w - 180

        for risk, count in risk_counts.items():
            pct = count / total
            # Label
            p.setFont(QFont("Inter", 9))
            p.setPen(QColor(Palette.TEXT_SECONDARY))
            p.drawText(QRectF(10, y, 80, 28), Qt.AlignRight | Qt.AlignVCenter, risk.upper())

            # Bar
            bar_w = bar_area_w * pct
            p.setPen(Qt.NoPen)
            p.setBrush(risk_colors.get(risk, QColor("#555")))
            p.drawRoundedRect(QRectF(96, y + 4, bar_w, 20), 3, 3)

            # Count
            p.setFont(QFont("JetBrains Mono", 9, QFont.Bold))
            p.setPen(QColor(Palette.TEXT_PRIMARY))
            p.drawText(QRectF(96 + bar_area_w + 6, y, 60, 28),
                       Qt.AlignLeft | Qt.AlignVCenter, f"{count} ({pct:.0%})")
            y += 36

        p.end()


class _DurationTimelineWidget(QWidget):
    """Gantt-like timeline showing step durations."""

    def __init__(self, route: Route, parent=None):
        super().__init__(parent)
        self.route = route
        self.setMinimumHeight(max(250, len(route.steps) * 28 + 60))
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setRenderHint(QPainter.TextAntialiasing, True)

        w, h = self.width(), self.height()
        steps = self.route.steps
        if not steps:
            p.end()
            return

        max_dur = max((s.duration_minutes for s in steps), default=1) or 1
        label_w = 160
        bar_area = w - label_w - 50
        bar_h = min(22, max(14, (h - 50) // len(steps) - 4))
        y = 10

        title_font = QFont("Inter", 10, QFont.Bold)
        p.setFont(title_font)
        p.setPen(QColor(Palette.GOLD_BRIGHT))
        p.drawText(10, y + 12, "Duration Timeline (minutes)")
        y += 28

        kind_colors = {
            "action": QColor("#5A7FA8"), "decision": QColor("#8B6FC0"),
            "milestone": QColor("#C9A84C"), "wait": QColor("#6B8FA3"),
            "checkpoint": QColor("#4A9A5A"), "research": QColor("#4A8AB0"),
            "review": QColor("#A06040"), "deliver": QColor("#5A8A5A"),
            "collaborate": QColor("#8A6A9A"),
        }

        for step in steps:
            # Label
            p.setFont(QFont("Inter", 8))
            p.setPen(QColor(Palette.TEXT_SECONDARY))
            label = step.title[:18] + "…" if len(step.title) > 18 else step.title
            p.drawText(QRectF(4, y, label_w - 8, bar_h), Qt.AlignRight | Qt.AlignVCenter, label)

            # Duration bar
            dur_pct = step.duration_minutes / max_dur
            bar_w = max(4, bar_area * dur_pct)
            color = kind_colors.get(step.kind, QColor("#5A7FA8"))

            p.setPen(Qt.NoPen)
            grad = QLinearGradient(label_w, 0, label_w + bar_w, 0)
            grad.setColorAt(0, color)
            grad.setColorAt(1, color.lighter(130))
            p.setBrush(QBrush(grad))
            p.drawRoundedRect(QRectF(label_w, y + 2, bar_w, bar_h - 4), 3, 3)

            # Duration text
            p.setFont(QFont("JetBrains Mono", 8))
            p.setPen(QColor(Palette.TEXT_TERTIARY))
            p.drawText(QRectF(label_w + bar_w + 4, y, 50, bar_h),
                       Qt.AlignLeft | Qt.AlignVCenter, f"{step.duration_minutes}m")

            y += bar_h + 4

        p.end()


class _KindDonutWidget(QWidget):
    """Donut chart showing step kind distribution."""

    def __init__(self, route: Route, parent=None):
        super().__init__(parent)
        self.route = route
        self.setFixedSize(260, 260)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)

        steps = self.route.steps
        if not steps:
            p.end()
            return

        kind_counts = {}
        for s in steps:
            kind_counts[s.kind] = kind_counts.get(s.kind, 0) + 1

        kind_colors = {
            "action": "#5A7FA8", "decision": "#8B6FC0", "milestone": "#C9A84C",
            "wait": "#6B8FA3", "checkpoint": "#4A9A5A", "research": "#4A8AB0",
            "review": "#A06040", "deliver": "#5A8A5A", "collaborate": "#8A6A9A",
        }

        total = len(steps)
        cx, cy, r = 130, 125, 90
        start_angle = 0

        for kind, count in kind_counts.items():
            span = int(360 * 16 * count / total)
            color = QColor(kind_colors.get(kind, "#555"))
            p.setPen(Qt.NoPen)
            p.setBrush(color)
            p.drawPie(QRectF(cx - r, cy - r, r * 2, r * 2), start_angle, span)
            start_angle += span

        # Center hole
        p.setBrush(QColor(Palette.BG_PRIMARY))
        p.drawEllipse(QRectF(cx - 55, cy - 55, 110, 110))

        # Center text
        p.setFont(QFont("Inter", 18, QFont.Bold))
        p.setPen(QColor(Palette.GOLD_BRIGHT))
        p.drawText(QRectF(cx - 50, cy - 18, 100, 24), Qt.AlignCenter, str(total))
        p.setFont(QFont("Inter", 8))
        p.setPen(QColor(Palette.TEXT_TERTIARY))
        p.drawText(QRectF(cx - 50, cy + 6, 100, 18), Qt.AlignCenter, "steps")

        # Legend
        y = 8
        for kind, count in kind_counts.items():
            color = QColor(kind_colors.get(kind, "#555"))
            p.setPen(Qt.NoPen)
            p.setBrush(color)
            p.drawRoundedRect(QRectF(210, y, 10, 10), 2, 2)
            p.setFont(QFont("Inter", 8))
            p.setPen(QColor(Palette.TEXT_SECONDARY))
            p.drawText(QRectF(224, y - 2, 40, 14), Qt.AlignLeft | Qt.AlignVCenter,
                       f"{kind[:4]} {count}")
            y += 16

        p.end()


class _BranchFlowWidget(QWidget):
    """Visual diagram showing branch flow with node counts per branch."""

    def __init__(self, route: Route, parent=None):
        super().__init__(parent)
        self.route = route
        self.setFixedHeight(200)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setRenderHint(QPainter.TextAntialiasing, True)

        w, h = self.width(), self.height()
        steps = self.route.steps
        if not steps:
            p.end()
            return

        # Count per branch
        branch_counts = {}
        for s in steps:
            branch_counts[s.branch] = branch_counts.get(s.branch, 0) + 1

        title_font = QFont("Inter", 10, QFont.Bold)
        p.setFont(title_font)
        p.setPen(QColor(Palette.GOLD_BRIGHT))
        p.drawText(10, 18, "Branch Distribution")

        branch_colors = {
            "main": QColor("#D4AF37"),
            "alt-1": QColor("#5A7FA8"),
            "alt-2": QColor("#8B6FC0"),
            "fallback-1": QColor("#A87A4A"),
            "fallback-2": QColor("#A85A5A"),
            "tasks": QColor("#6B8FA3"),
        }

        n = len(branch_counts)
        if n == 0:
            p.end()
            return

        col_w = min(180, (w - 40) // n)
        x = 20
        y_base = h - 40
        max_count = max(branch_counts.values()) if branch_counts else 1
        max_bar_h = h - 80

        for branch, count in branch_counts.items():
            bar_h = max(10, int(max_bar_h * count / max_count))
            color = branch_colors.get(branch, QColor("#555"))

            # Bar
            p.setPen(Qt.NoPen)
            grad = QLinearGradient(x, y_base - bar_h, x, y_base)
            grad.setColorAt(0, color.lighter(120))
            grad.setColorAt(1, color)
            p.setBrush(QBrush(grad))
            p.drawRoundedRect(QRectF(x, y_base - bar_h, col_w - 10, bar_h), 4, 4)

            # Count on top
            p.setFont(QFont("JetBrains Mono", 10, QFont.Bold))
            p.setPen(QColor(Palette.TEXT_PRIMARY))
            p.drawText(QRectF(x, y_base - bar_h - 18, col_w - 10, 16),
                       Qt.AlignCenter, str(count))

            # Label below
            p.setFont(QFont("Inter", 8))
            p.setPen(QColor(Palette.TEXT_SECONDARY))
            p.drawText(QRectF(x, y_base + 4, col_w - 10, 16),
                       Qt.AlignCenter, branch[:10])

            x += col_w

        p.end()


# ──────────────────────────────────────────────────────────────────────
#  Graphs Landing Page
# ──────────────────────────────────────────────────────────────────────

class _GraphsLanding(QFrame):
    """Landing page where the user selects a workflow to visualize."""

    workflowSelected = Signal(object)  # Route

    def __init__(self, journal_store: JournalStore, parent=None):
        super().__init__(parent)
        self.journal_store = journal_store
        self._build_ui()

    def _build_ui(self):
        self.setStyleSheet(f"background-color: {Palette.BG_PRIMARY};")

        outer = QVBoxLayout(self)
        outer.setContentsMargins(40, 40, 40, 40)
        outer.setSpacing(20)

        # Title
        title = QLabel("📊  WORKFLOW ANALYTICS")
        title.setStyleSheet(
            f"color: {Palette.GOLD_BRIGHT}; font-size: 24px; "
            f"font-weight: bold; letter-spacing: 3px; background: transparent;"
        )
        outer.addWidget(title)

        subtitle = QLabel("Select a workflow to visualize, or import one from a file.")
        subtitle.setStyleSheet(
            f"color: {Palette.TEXT_SECONDARY}; font-size: 13px; background: transparent;"
        )
        outer.addWidget(subtitle)

        # Import/Export buttons
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)

        self._import_csv_btn = QPushButton("📁  Import CSV")
        self._import_csv_btn.setStyleSheet(self._btn_style())
        self._import_csv_btn.clicked.connect(self._on_import_csv)
        btn_row.addWidget(self._import_csv_btn)

        self._import_excel_btn = QPushButton("📊  Import Excel")
        self._import_excel_btn.setStyleSheet(self._btn_style())
        self._import_excel_btn.clicked.connect(self._on_import_excel)
        btn_row.addWidget(self._import_excel_btn)

        btn_row.addStretch()
        outer.addLayout(btn_row)

        # Separator
        sep = QFrame()
        sep.setFixedHeight(1)
        sep.setStyleSheet(f"background-color: {Palette.BORDER_NORMAL};")
        outer.addWidget(sep)

        # Workflow list header
        list_header = QLabel("SAVED WORKFLOWS")
        list_header.setStyleSheet(
            f"color: {Palette.TEXT_TERTIARY}; font-size: 10px; "
            f"font-weight: bold; letter-spacing: 2px; background: transparent;"
        )
        outer.addWidget(list_header)

        # Scroll area for workflows
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"""
            QScrollArea {{ background: {Palette.BG_SECONDARY}; border: 1px solid {Palette.BORDER_SUBTLE}; border-radius: 8px; }}
            QScrollBar:vertical {{ background: {Palette.BG_TERTIARY}; width: 8px; }}
            QScrollBar::handle:vertical {{ background: {Palette.BORDER_NORMAL}; border-radius: 4px; }}
        """)
        scroll_content = QWidget()
        self._workflow_layout = QVBoxLayout(scroll_content)
        self._workflow_layout.setContentsMargins(12, 12, 12, 12)
        self._workflow_layout.setSpacing(8)
        self._workflow_layout.addStretch()
        scroll.setWidget(scroll_content)
        outer.addWidget(scroll, stretch=1)

        self.refresh_workflows()

    def _btn_style(self) -> str:
        return f"""
            QPushButton {{
                background-color: {Palette.BG_TERTIARY};
                color: {Palette.GOLD_BRIGHT};
                border: 1px solid {Palette.BORDER_GOLD};
                border-radius: 6px;
                padding: 10px 20px;
                font-size: 12px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {Palette.BG_SELECTED};
                border: 1px solid {Palette.GOLD_BRIGHT};
            }}
        """

    def _workflow_card_style(self) -> str:
        return f"""
            QFrame {{
                background-color: {Palette.BG_ELEVATED};
                border: 1px solid {Palette.BORDER_NORMAL};
                border-left: 3px solid {Palette.GOLD_PRIMARY};
                border-radius: 6px;
                padding: 4px;
            }}
            QFrame:hover {{
                border-color: {Palette.GOLD_PRIMARY};
                background-color: {Palette.BG_HOVER};
            }}
        """

    def refresh_workflows(self):
        # Clear existing cards
        while self._workflow_layout.count() > 1:
            item = self._workflow_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        # Add workflow cards from journal
        entries = self.journal_store.all()
        for entry in entries:
            if entry.route is None:
                continue
            card = QFrame()
            card.setStyleSheet(self._workflow_card_style())
            card.setCursor(Qt.PointingHandCursor)
            cl = QVBoxLayout(card)
            cl.setContentsMargins(12, 8, 12, 8)
            cl.setSpacing(4)

            goal_label = QLabel(f"<b>{entry.user_goal[:60]}</b>")
            goal_label.setStyleSheet(f"color: {Palette.TEXT_PRIMARY}; background: transparent; font-size: 12px;")
            goal_label.setWordWrap(True)
            cl.addWidget(goal_label)

            meta = QLabel(
                f"{len(entry.route.steps)} steps · {entry.route.total_duration_minutes}m · "
                f"{entry.route.overall_success_probability:.0%} success · "
                f"{entry.timestamp[:10]}"
            )
            meta.setStyleSheet(f"color: {Palette.TEXT_TERTIARY}; background: transparent; font-size: 10px;")
            cl.addWidget(meta)

            # Export buttons
            btn_row = QHBoxLayout()
            btn_row.setSpacing(6)

            export_csv_btn = QPushButton("CSV")
            export_csv_btn.setFixedSize(50, 22)
            export_csv_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {Palette.BG_TERTIARY};
                    color: {Palette.TEXT_TERTIARY};
                    border: 1px solid {Palette.BORDER_NORMAL};
                    border-radius: 3px;
                    font-size: 9px;
                }}
                QPushButton:hover {{
                    color: {Palette.GOLD_BRIGHT};
                    border-color: {Palette.GOLD_PRIMARY};
                }}
            """)
            route_ref = entry.route
            export_csv_btn.clicked.connect(lambda _, r=route_ref: self._export_csv(r))
            btn_row.addWidget(export_csv_btn)

            export_xl_btn = QPushButton("Excel")
            export_xl_btn.setFixedSize(50, 22)
            export_xl_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {Palette.BG_TERTIARY};
                    color: {Palette.TEXT_TERTIARY};
                    border: 1px solid {Palette.BORDER_NORMAL};
                    border-radius: 3px;
                    font-size: 9px;
                }}
                QPushButton:hover {{
                    color: {Palette.GOLD_BRIGHT};
                    border-color: {Palette.GOLD_PRIMARY};
                }}
            """)
            export_xl_btn.clicked.connect(lambda _, r=route_ref: self._export_excel(r))
            btn_row.addWidget(export_xl_btn)
            btn_row.addStretch()
            cl.addLayout(btn_row)

            # Click to select
            card.mousePressEvent = lambda e, r=entry.route: self.workflowSelected.emit(r)
            self._workflow_layout.insertWidget(self._workflow_layout.count() - 1, card)

    def _on_import_csv(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Workflow from CSV", "",
            "CSV files (*.csv);;All files (*)"
        )
        if path:
            try:
                route = import_route_csv(path)
                self.workflowSelected.emit(route)
            except Exception as e:
                QMessageBox.warning(self, "Import Failed", str(e))

    def _on_import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Workflow from Excel", "",
            "Excel files (*.xlsx);;All files (*)"
        )
        if path:
            try:
                route = import_route_excel(path)
                self.workflowSelected.emit(route)
            except Exception as e:
                QMessageBox.warning(self, "Import Failed", str(e))

    def _export_csv(self, route: Route):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Workflow as CSV", "workflow.csv",
            "CSV files (*.csv);;All files (*)"
        )
        if path:
            try:
                export_route_csv(route, path)
                QMessageBox.information(self, "Exported", f"Saved to {path}")
            except Exception as e:
                QMessageBox.warning(self, "Export Failed", str(e))

    def _export_excel(self, route: Route):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Workflow as Excel", "workflow.xlsx",
            "Excel files (*.xlsx);;All files (*)"
        )
        if path:
            try:
                export_route_excel(route, path)
                QMessageBox.information(self, "Exported", f"Saved to {path}")
            except Exception as e:
                QMessageBox.warning(self, "Export Failed", str(e))


# ──────────────────────────────────────────────────────────────────────
#  Graphs Workspace (shows charts for a selected route)
# ──────────────────────────────────────────────────────────────────────

class _GraphsWorkspace(QFrame):
    """Workspace showing all chart visualizations for a selected route."""

    backRequested = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._route: Optional[Route] = None
        self._build_ui()

    def _build_ui(self):
        self.setStyleSheet(f"background-color: {Palette.BG_PRIMARY};")

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # Top bar
        top = QFrame()
        top.setFixedHeight(44)
        top.setStyleSheet(f"background-color: {Palette.BG_SECONDARY}; border-bottom: 1px solid {Palette.BORDER_SUBTLE};")
        top_layout = QHBoxLayout(top)
        top_layout.setContentsMargins(12, 6, 12, 6)

        back_btn = QPushButton("← Back")
        back_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {Palette.TEXT_TERTIARY};
                border: 1px solid {Palette.BORDER_NORMAL}; border-radius: 3px;
                padding: 4px 12px; font-size: 11px;
            }}
            QPushButton:hover {{
                color: {Palette.GOLD_BRIGHT}; border-color: {Palette.GOLD_PRIMARY};
            }}
        """)
        back_btn.clicked.connect(self.backRequested.emit)
        top_layout.addWidget(back_btn)

        self._route_title = QLabel("")
        self._route_title.setStyleSheet(
            f"color: {Palette.GOLD_BRIGHT}; font-size: 12px; font-weight: bold; background: transparent;"
        )
        top_layout.addWidget(self._route_title)
        top_layout.addStretch()

        # Export buttons in top bar
        self._export_csv_btn = QToolButton()
        self._export_csv_btn.setText("📁 CSV")
        self._export_csv_btn.setStyleSheet(f"""
            QToolButton {{
                background-color: {Palette.BG_TERTIARY}; color: {Palette.TEXT_TERTIARY};
                border: 1px solid {Palette.BORDER_NORMAL}; border-radius: 3px;
                padding: 4px 10px; font-size: 10px;
            }}
            QToolButton:hover {{
                color: {Palette.GOLD_BRIGHT}; border-color: {Palette.GOLD_PRIMARY};
            }}
        """)
        self._export_csv_btn.clicked.connect(self._on_export_csv)
        top_layout.addWidget(self._export_csv_btn)

        self._export_xl_btn = QToolButton()
        self._export_xl_btn.setText("📊 Excel")
        self._export_xl_btn.setStyleSheet(f"""
            QToolButton {{
                background-color: {Palette.BG_TERTIARY}; color: {Palette.TEXT_TERTIARY};
                border: 1px solid {Palette.BORDER_NORMAL}; border-radius: 3px;
                padding: 4px 10px; font-size: 10px;
            }}
            QToolButton:hover {{
                color: {Palette.GOLD_BRIGHT}; border-color: {Palette.GOLD_PRIMARY};
            }}
        """)
        self._export_xl_btn.clicked.connect(self._on_export_excel)
        top_layout.addWidget(self._export_xl_btn)

        outer.addWidget(top)

        # Scroll area for charts
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"""
            QScrollArea {{ background: {Palette.BG_PRIMARY}; border: none; }}
            QScrollBar:vertical {{ background: {Palette.BG_TERTIARY}; width: 8px; }}
            QScrollBar::handle:vertical {{ background: {Palette.BORDER_NORMAL}; border-radius: 4px; }}
        """)

        self._charts_container = QWidget()
        self._charts_layout = QVBoxLayout(self._charts_container)
        self._charts_layout.setContentsMargins(20, 16, 20, 20)
        self._charts_layout.setSpacing(16)

        scroll.setWidget(self._charts_container)
        outer.addWidget(scroll, stretch=1)

    def set_route(self, route: Route):
        self._route = route
        self._route_title.setText(
            f"📊  {route.goal[:50]}  ·  {len(route.steps)} steps  ·  {route.total_duration_minutes}m"
        )
        self._rebuild_charts()

    def _rebuild_charts(self):
        # Clear existing charts
        while self._charts_layout.count():
            item = self._charts_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        if self._route is None:
            return

        route = self._route

        # Stats summary row
        stats_row = QHBoxLayout()
        stats_row.setSpacing(12)

        for label, value, color in [
            ("Steps", str(len(route.steps)), Palette.GOLD_BRIGHT),
            ("Duration", f"{route.total_duration_minutes}m", Palette.TEXT_PRIMARY),
            ("Success", f"{route.overall_success_probability:.0%}", "#5A8A5A" if route.overall_success_probability > 0.7 else "#D4AF37"),
            ("Edges", str(len(route.edges)), Palette.TEXT_SECONDARY),
            ("Insights", str(len(route.insights)), Palette.TEXT_TERTIARY),
            ("Branches", str(len(set(s.branch for s in route.steps))), Palette.TEXT_SECONDARY),
        ]:
            stat = QFrame()
            stat.setStyleSheet(f"""
                QFrame {{
                    background-color: {Palette.BG_ELEVATED};
                    border: 1px solid {Palette.BORDER_SUBTLE};
                    border-radius: 8px;
                    padding: 8px;
                }}
            """)
            sl = QVBoxLayout(stat)
            sl.setContentsMargins(12, 8, 12, 8)
            sl.setSpacing(2)
            v = QLabel(value)
            v.setStyleSheet(f"color: {color}; font-size: 20px; font-weight: bold; background: transparent;")
            v.setAlignment(Qt.AlignCenter)
            sl.addWidget(v)
            l = QLabel(label)
            l.setStyleSheet(f"color: {Palette.TEXT_TERTIARY}; font-size: 9px; background: transparent; letter-spacing: 1px;")
            l.setAlignment(Qt.AlignCenter)
            sl.addWidget(l)
            stats_row.addWidget(stat)

        self._charts_layout.addLayout(stats_row)

        # Chart row 1: Success probability bars + Kind donut
        row1 = QHBoxLayout()
        row1.setSpacing(16)

        bar_card = self._chart_card("Success Probability", _BarChartWidget(route))
        row1.addWidget(bar_card, stretch=3)

        donut_card = self._chart_card("Step Kinds", _KindDonutWidget(route))
        row1.addWidget(donut_card, stretch=1)

        self._charts_layout.addLayout(row1)

        # Chart row 2: Duration timeline
        dur_card = self._chart_card("Duration Timeline", _DurationTimelineWidget(route))
        self._charts_layout.addWidget(dur_card)

        # Chart row 3: Risk heatmap + Branch flow
        row3 = QHBoxLayout()
        row3.setSpacing(16)

        risk_card = self._chart_card("Risk Levels", _RiskHeatmapWidget(route))
        row3.addWidget(risk_card, stretch=1)

        branch_card = self._chart_card("Branch Distribution", _BranchFlowWidget(route))
        row3.addWidget(branch_card, stretch=1)

        self._charts_layout.addLayout(row3)

    def _chart_card(self, title: str, chart_widget: QWidget) -> QFrame:
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {Palette.BG_SECONDARY};
                border: 1px solid {Palette.BORDER_SUBTLE};
                border-radius: 10px;
            }}
        """)
        cl = QVBoxLayout(card)
        cl.setContentsMargins(0, 0, 0, 0)
        cl.setSpacing(0)
        cl.addWidget(chart_widget)
        return card

    def _on_export_csv(self):
        if not self._route:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Workflow as CSV", "workflow.csv",
            "CSV files (*.csv);;All files (*)"
        )
        if path:
            try:
                export_route_csv(self._route, path)
                QMessageBox.information(self, "Exported", f"Saved to {path}")
            except Exception as e:
                QMessageBox.warning(self, "Export Failed", str(e))

    def _on_export_excel(self):
        if not self._route:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Workflow as Excel", "workflow.xlsx",
            "Excel files (*.xlsx);;All files (*)"
        )
        if path:
            try:
                export_route_excel(self._route, path)
                QMessageBox.information(self, "Exported", f"Saved to {path}")
            except Exception as e:
                QMessageBox.warning(self, "Export Failed", str(e))


# ──────────────────────────────────────────────────────────────────────
#  Main GraphsView (stacked: landing + workspace)
# ──────────────────────────────────────────────────────────────────────

class GraphsView(QWidget):
    """
    Full-page visual analytics for RASK! routes/workflows.

    Page 0: Landing — select workflow from journal or import CSV/Excel
    Page 1: Workspace — charts for the selected route
    """

    routeSelected = Signal(object)  # Route — emitted when user picks a route

    def __init__(self, journal_store: JournalStore, parent=None):
        super().__init__(parent)
        self.journal_store = journal_store
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self._stack = QStackedWidget()

        # Page 0: Landing
        self._landing = _GraphsLanding(self.journal_store)
        self._landing.workflowSelected.connect(self._on_workflow_selected)
        self._stack.addWidget(self._landing)

        # Page 1: Workspace
        self._workspace = _GraphsWorkspace()
        self._workspace.backRequested.connect(self._show_landing)
        self._stack.addWidget(self._workspace)

        layout.addWidget(self._stack)

    def _on_workflow_selected(self, route: Route):
        self._workspace.set_route(route)
        self._stack.setCurrentIndex(1)
        self.routeSelected.emit(route)

    def _show_landing(self):
        self._landing.refresh_workflows()
        self._stack.setCurrentIndex(0)

    def set_route(self, route: Route):
        """Programmatically set a route and show charts."""
        self._workspace.set_route(route)
        self._stack.setCurrentIndex(1)

    def refresh(self):
        """Refresh the landing page workflow list."""
        self._landing.refresh_workflows()
