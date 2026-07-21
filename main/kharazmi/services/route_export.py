"""
Route export utilities — CSV, Excel (xlsx), HTML.

Exports the steps and edges of a Route to various file formats.
"""
from __future__ import annotations

import csv
import io
import html
from pathlib import Path
from typing import Optional

from ..ai.ai_service import Route, RouteStep, RouteEdge


def export_route_csv(route: Route, path: str | Path) -> None:
    """Export route steps to CSV file."""
    path = Path(path)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "#", "Title", "Description", "Kind", "Branch",
            "Duration (min)", "Success %", "Risk Level",
            "Location", "Cost Estimate", "Fallback",
            "Depends On", "Sub-Goals"
        ])
        for i, step in enumerate(route.steps, 1):
            writer.writerow([
                i,
                step.title,
                step.description,
                step.kind,
                step.branch,
                step.duration_minutes,
                f"{step.success_probability:.0%}",
                step.risk_level,
                step.location,
                step.cost_estimate,
                step.fallback,
                ", ".join(step.depends_on),
                ", ".join(step.sub_goals),
            ])


def export_route_xlsx(route: Route, path: str | Path) -> None:
    """Export route steps to Excel (xlsx) file using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback: write as CSV with .xlsx extension warning
        export_route_csv(route, str(path).replace(".xlsx", ".csv"))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Route Steps"

    # Header styling
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D2D35", end_color="2D2D35", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="555555"),
        right=Side(style="thin", color="555555"),
        top=Side(style="thin", color="555555"),
        bottom=Side(style="thin", color="555555"),
    )

    headers = [
        "#", "Title", "Description", "Kind", "Branch",
        "Duration (min)", "Success %", "Risk Level",
        "Location", "Cost Estimate", "Fallback",
        "Depends On", "Sub-Goals"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    data_font = Font(name="Segoe UI", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    # Risk color mapping
    risk_fills = {
        "low": PatternFill(start_color="1A3A1A", end_color="1A3A1A", fill_type="solid"),
        "medium": PatternFill(start_color="3A3A1A", end_color="3A3A1A", fill_type="solid"),
        "high": PatternFill(start_color="3A1A1A", end_color="3A1A1A", fill_type="solid"),
    }

    for i, step in enumerate(route.steps, 2):
        row_data = [
            i - 1,
            step.title,
            step.description,
            step.kind,
            step.branch,
            step.duration_minutes,
            step.success_probability,
            step.risk_level,
            step.location,
            step.cost_estimate,
            step.fallback,
            ", ".join(step.depends_on),
            ", ".join(step.sub_goals),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            # Color risk column
            if col == 8:  # Risk Level
                cell.fill = risk_fills.get(str(val).lower(), PatternFill())

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 40)

    # Add Edges sheet
    ws2 = wb.create_sheet("Edges")
    edge_headers = ["Source", "Target", "Kind", "Label"]
    for col, header in enumerate(edge_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, edge in enumerate(route.edges, 2):
        ws2.cell(row=i, column=1, value=edge.source_id).font = data_font
        ws2.cell(row=i, column=2, value=edge.target_id).font = data_font
        ws2.cell(row=i, column=3, value=edge.kind).font = data_font
        ws2.cell(row=i, column=4, value=edge.label).font = data_font

    # Summary sheet
    ws3 = wb.create_sheet("Summary")
    summary_data = [
        ("Goal", route.goal),
        ("Overall Success", f"{route.overall_success_probability:.0%}"),
        ("Total Duration", f"{route.total_duration_minutes} min"),
        ("Steps", len(route.steps)),
        ("Edges", len(route.edges)),
        ("Layout Style", route.layout_style),
        ("Summary", route.summary),
    ]
    for i, (key, val) in enumerate(summary_data, 1):
        cell_k = ws3.cell(row=i, column=1, value=key)
        cell_k.font = Font(name="Segoe UI", bold=True, size=11)
        cell_v = ws3.cell(row=i, column=2, value=val)
        cell_v.font = data_font
        cell_v.alignment = Alignment(wrap_text=True)
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 60

    wb.save(path)


def export_route_html(route: Route, path: str | Path) -> None:
    """Export route as a self-contained HTML page with dark theme."""
    path = Path(path)

    rows_html = ""
    for i, step in enumerate(route.steps, 1):
        risk_colors = {"low": "#2d5a2d", "medium": "#5a5a2d", "high": "#5a2d2d"}
        risk_bg = risk_colors.get(step.risk_level.lower(), "#2d2d35")
        pct = step.success_probability * 100
        pct_color = "#4CAF50" if pct > 70 else ("#FFC107" if pct > 40 else "#F44336")

        rows_html += f"""
        <tr>
            <td>{i}</td>
            <td><strong>{html.escape(step.title)}</strong></td>
            <td>{html.escape(step.description)}</td>
            <td><span style="background:{risk_bg};padding:2px 8px;border-radius:4px;">{html.escape(step.kind)}</span></td>
            <td>{html.escape(step.branch)}</td>
            <td>{step.duration_minutes}</td>
            <td style="color:{pct_color};font-weight:bold;">{pct:.0f}%</td>
            <td><span style="background:{risk_bg};padding:2px 8px;border-radius:4px;">{html.escape(step.risk_level)}</span></td>
            <td>{html.escape(step.location)}</td>
            <td>{html.escape(step.fallback)}</td>
        </tr>"""

    page = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RASK! — {html.escape(route.goal)}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:'Segoe UI',Tahoma,sans-serif; background:#0a0a0e; color:#e0e0e0; padding:40px; }}
  h1 {{ color:#D4AF37; font-size:28px; margin-bottom:8px; }}
  h2 {{ color:#D4AF37; font-size:20px; margin:32px 0 16px; border-bottom:1px solid #333; padding-bottom:8px; }}
  .summary {{ background:#141418; border:1px solid #2a2a30; border-radius:12px; padding:20px; margin:16px 0; }}
  .summary-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:16px; margin-top:16px; }}
  .stat {{ background:#1a1a20; border-radius:8px; padding:16px; text-align:center; }}
  .stat .num {{ font-size:28px; font-weight:bold; color:#D4AF37; }}
  .stat .lbl {{ font-size:12px; color:#888; margin-top:4px; }}
  table {{ width:100%; border-collapse:collapse; margin:16px 0; }}
  th {{ background:#1a1a22; color:#D4AF37; padding:12px 16px; text-align:right; font-size:12px; letter-spacing:1px; }}
  td {{ padding:10px 16px; border-bottom:1px solid #1a1a22; font-size:13px; }}
  tr:hover {{ background:#141418; }}
  .badge {{ background:#1a1a22; color:#D4AF37; padding:2px 8px; border-radius:4px; font-size:11px; }}
  footer {{ margin-top:40px; padding-top:16px; border-top:1px solid #2a2a30; color:#555; font-size:11px; }}
</style>
</head>
<body>
<h1>&#10022; RASK! Route Export</h1>
<p style="color:#888;margin-bottom:24px;">{html.escape(route.goal)}</p>

<div class="summary">
  <div class="summary-grid">
    <div class="stat"><div class="num">{len(route.steps)}</div><div class="lbl">Steps</div></div>
    <div class="stat"><div class="num">{route.overall_success_probability:.0%}</div><div class="lbl">Success Probability</div></div>
    <div class="stat"><div class="num">{route.total_duration_minutes} min</div><div class="lbl">Total Duration</div></div>
    <div class="stat"><div class="num">{len(route.edges)}</div><div class="lbl">Connections</div></div>
  </div>
</div>

{f'<p style="color:#aaa;margin:16px 0;">{html.escape(route.summary)}</p>' if route.summary else ''}

<h2>Steps</h2>
<table>
<thead>
<tr>
  <th>#</th><th>Title</th><th>Description</th><th>Kind</th><th>Branch</th>
  <th>Duration</th><th>Success</th><th>Risk</th><th>Location</th><th>Fallback</th>
</tr>
</thead>
<tbody>
{rows_html}
</tbody>
</table>

<footer>Generated by RASK! v3.0 &mdash; Kharazmi AI Planning System</footer>
</body>
</html>"""

    path.write_text(page, encoding="utf-8")
